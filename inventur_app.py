#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
INVENTUR-PROGRAMM V2 FÜR FORBO MOVEMENT SYSTEMS
===============================================

Desktop-Anwendung für Lagerpflege/Inventur mit Barcodescanner-Integration
Unterstützt Rollen UND Granulate mit separaten Excel-Dateien
Entwickelt für Windows 11, Python 3.11+

Autor: Cascade AI
Datum: Dezember 2024
Version: 2.0
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import sys
from pathlib import Path
import json
import logging

class InventurApp:
    def __init__(self):
        """Initialisiert die Inventur-Anwendung"""
        self.root = tk.Tk()
        
        # EXE-kompatible Pfade definieren (ZUERST!)
        self.base_dir = Path(self.get_base_path())
        self.data_dir = self.base_dir / 'data'
        self.config_dir = self.base_dir / 'config'
        self.arbeitstabelle_path = self.data_dir / 'Arbeitstabelle.xlsx'
        self.inventur_rollen_path = self.data_dir / 'Inventur_Rollen.xlsx'
        self.inventur_granulat_path = self.data_dir / 'Inventur_Granulat.xlsx'
        
        # Erstelle Verzeichnisse falls nicht vorhanden
        self.data_dir.mkdir(exist_ok=True)
        self.config_dir.mkdir(exist_ok=True)
        (self.data_dir / 'backups').mkdir(exist_ok=True)
        
        # Jetzt Logging setup (nachdem Pfade definiert sind)
        self.setup_logging()
        
        self.load_config()
        self.init_data()
        self.setup_ui()
        self.load_existing_inventur()
        self.bind_shortcuts()
        
    def get_base_path(self):
        """Gibt den Basispfad zurück - funktioniert sowohl für .py als auch .exe"""
        if getattr(sys, 'frozen', False):
            # Läuft als EXE (PyInstaller)
            return os.path.dirname(sys.executable)
        else:
            # Läuft als Python-Skript
            return os.path.dirname(os.path.abspath(__file__))
    
    def setup_logging(self):
        """Konfiguriert das Logging-System"""
        log_file = self.data_dir / 'inventur.log'
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_file, encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
        self.logger.info("Inventur-Programm V2 gestartet")
    
    def load_config(self):
        """Lädt die Konfigurationsdatei"""
        config_path = self.config_dir / 'settings.json'
        default_config = {
            "auto_save": True,
            "schriftgroesse": 12,
            "farbe_gefunden": "#E8F5E8",
            "farbe_nicht_gefunden": "#FFF2CC",
            "farbe_fehler": "#FFE6E6",
            "farbe_rolle_bg": "#E3F2FD",
            "farbe_rolle_text": "#1976D2",
            "farbe_granulat_bg": "#FFF9C4",
            "farbe_granulat_text": "#F57F17",
            "vollbild": True
        }
        
        try:
            if config_path.exists():
                with open(config_path, 'r', encoding='utf-8-sig') as f:
                    self.config = json.load(f)
            else:
                self.config = default_config
                self.save_config()
        except Exception as e:
            self.config = default_config
            self.logger.error(f"Fehler beim Laden der Konfiguration: {e}")
    
    def save_config(self):
        """Speichert die Konfiguration"""
        try:
            config_path = self.config_dir / 'settings.json'
            with open(config_path, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=2, ensure_ascii=False)
        except Exception as e:
            self.logger.error(f"Fehler beim Speichern der Konfiguration: {e}")
    
    def init_data(self):
        """Initialisiert die Datenstrukturen"""
        # Separate DataFrames für Rollen und Granulat
        self.df_rollen = None
        self.df_granulate = None
        
        # Separate Listen für Inventur-Daten
        self.inventur_rollen_data = []
        self.inventur_granulat_data = []
        self.nicht_gefunden_rollen_data = []
        self.nicht_gefunden_granulat_data = []
        
        self.current_scan = None
        self.current_type = None  # 'ROLLE' oder 'GRANULAT'
        self.undo_stack = []
        
        # Lade Arbeitstabelle
        self.load_arbeitstabelle()
    
    def load_arbeitstabelle(self):
        """Lädt die Arbeitstabelle mit zwei Tabellenblättern (Rollen und Granulate)"""
        try:
            if self.arbeitstabelle_path.exists():
                # Prüfe ob beide Tabellenblätter vorhanden sind
                try:
                    excel_file = pd.ExcelFile(self.arbeitstabelle_path)
                    available_sheets = excel_file.sheet_names
                    
                    if 'Rollen' not in available_sheets or 'Granulate' not in available_sheets:
                        messagebox.showerror("Fehler", 
                            f"⚠️ FEHLER: Arbeitstabelle.xlsx muss zwei Tabellenblätter haben:\n"
                            f"- 'Rollen'\n"
                            f"- 'Granulate'\n\n"
                            f"Gefundene Blätter: {', '.join(available_sheets)}\n\n"
                            f"Bitte überprüfen Sie die Datei.")
                        sys.exit(1)
                    
                    # Lade Rollen-Tabellenblatt
                    self.df_rollen = pd.read_excel(self.arbeitstabelle_path, sheet_name='Rollen', dtype={'Charge': str})
                    if 'Charge' in self.df_rollen.columns:
                        self.df_rollen['Charge'] = self.df_rollen['Charge'].astype(str)
                    
                    # Lade Granulate-Tabellenblatt
                    self.df_granulate = pd.read_excel(self.arbeitstabelle_path, sheet_name='Granulate', dtype={'Charge': str})
                    if 'Charge' in self.df_granulate.columns:
                        self.df_granulate['Charge'] = self.df_granulate['Charge'].astype(str)
                    
                    # Umbenennen: "Materialnummer" → "Material" (für einheitliche Verarbeitung)
                    if 'Materialnummer' in self.df_granulate.columns:
                        self.df_granulate.rename(columns={'Materialnummer': 'Material'}, inplace=True)
                    
                    # Prüfe erforderliche Spalten für Rollen
                    required_rollen_columns = ['Charge', 'Material', 'Materialkurztext', 'Länge m', 'Breite mm', 'Frei verwendbar']
                    missing_rollen = [col for col in required_rollen_columns if col not in self.df_rollen.columns]
                    
                    # Prüfe erforderliche Spalten für Granulate
                    required_granulate_columns = ['Charge', 'Material', 'Materialkurztext', 'Frei verwendbar']
                    missing_granulate = [col for col in required_granulate_columns if col not in self.df_granulate.columns]
                    
                    if missing_rollen or missing_granulate:
                        error_msg = "Fehlende Spalten:\n"
                        if missing_rollen:
                            error_msg += f"Rollen: {', '.join(missing_rollen)}\n"
                        if missing_granulate:
                            error_msg += f"Granulate: {', '.join(missing_granulate)}"
                        messagebox.showerror("Fehler", error_msg)
                        sys.exit(1)
                    
                    rollen_count = len(self.df_rollen)
                    granulate_count = len(self.df_granulate)
                    total_count = rollen_count + granulate_count
                    
                    self.logger.info(f"Arbeitstabelle geladen: {rollen_count} Rollen, {granulate_count} Granulate, {total_count} gesamt")
                    
                except Exception as e:
                    messagebox.showerror("Fehler", f"Fehler beim Lesen der Excel-Datei:\n{e}")
                    sys.exit(1)
                    
            else:
                messagebox.showwarning("Warnung", 
                    f"Arbeitstabelle nicht gefunden!\n\n"
                    f"Bitte kopieren Sie die Arbeitstabelle.xlsx in:\n"
                    f"{self.arbeitstabelle_path.absolute()}\n\n"
                    f"Die Datei muss zwei Tabellenblätter enthalten:\n"
                    f"- 'Rollen'\n"
                    f"- 'Granulate'")
                
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Laden der Arbeitstabelle:\n{e}")
            self.logger.error(f"Fehler beim Laden der Arbeitstabelle: {e}")
    
    def setup_ui(self):
        """Erstellt die Benutzeroberfläche"""
        # Hauptfenster konfigurieren
        self.root.title("INVENTUR Forbo - Lagerverwaltung")
        self.root.geometry("1200x800")
        self.root.configure(bg='#f0f0f0')
        
        # Icon setzen (falls vorhanden)
        try:
            self.root.iconbitmap('config/forbo_icon.ico')
        except:
            pass
        
        # Vollbild-Modus - standardmäßig aktiviert für Inventur-Anwendung
        self.root.state('zoomed')  # Maximiert das Fenster
        # Optional: Echter Vollbild-Modus (ohne Taskleiste)
        # self.root.attributes('-fullscreen', True)
        
        # Hauptframe
        self.main_frame = ttk.Frame(self.root, padding="10")
        self.main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Grid-Konfiguration
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        self.main_frame.columnconfigure(0, weight=1)
        
        self.create_header()
        self.create_scan_section()
        self.create_current_scan_section()
        self.create_list_section()
        self.create_button_section()
        
        # Status-Bar
        self.create_status_bar()
        
        # Fokus auf Scan-Eingabefeld
        self.scan_entry.focus_set()
    
    def create_header(self):
        """Erstellt den Header-Bereich"""
        header_frame = ttk.Frame(self.main_frame)
        header_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 20))
        header_frame.columnconfigure(1, weight=1)
        
        # Logo-Platzhalter (links)
        logo_label = ttk.Label(header_frame, text="📦", font=("Arial", 24))
        logo_label.grid(row=0, column=0, padx=(0, 20))
        
        # Titel (mitte)
        title_label = ttk.Label(header_frame, 
                               text="INVENTUR Forbo - Lagerverwaltung",
                               font=("Arial", 18, "bold"),
                               foreground="#1f4e79")
        title_label.grid(row=0, column=1)
        
        # Info (rechts)
        if self.df_rollen is not None and self.df_granulate is not None:
            rollen_count = len(self.df_rollen)
            granulate_count = len(self.df_granulate)
            total_count = rollen_count + granulate_count
            info_text = f"DB: {rollen_count} 🔵 Rollen, {granulate_count} 🟨 Granulate ({total_count} gesamt)"
        else:
            info_text = "Keine Arbeitstabelle"
        
        info_label = ttk.Label(header_frame, text=info_text, font=("Arial", 10))
        info_label.grid(row=0, column=2, padx=(20, 0))
    
    def create_scan_section(self):
        """Erstellt den Barcode-Scan-Bereich"""
        # Scan-Frame
        scan_frame = ttk.LabelFrame(self.main_frame, text="🔍 BARCODE SCANNEN", padding="10")
        scan_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        scan_frame.columnconfigure(0, weight=1)
        
        # Scan-Eingabefeld
        self.scan_var = tk.StringVar()
        self.scan_entry = ttk.Entry(scan_frame, 
                                   textvariable=self.scan_var,
                                   font=("Arial", 14),
                                   width=50)
        self.scan_entry.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 10))
        
        # Scan-Button (falls Scanner kein Enter sendet)
        scan_button = ttk.Button(scan_frame, text="Scannen", command=self.process_scan)
        scan_button.grid(row=0, column=1)
        
        # Enter-Binding
        self.scan_entry.bind('<Return>', lambda e: self.process_scan())
        
        # Info-Label
        info_label = ttk.Label(scan_frame, 
                              text="Eingabefeld ist immer fokussiert. Barcode scannen oder eingeben + ENTER drücken.",
                              font=("Arial", 9),
                              foreground="gray")
        info_label.grid(row=1, column=0, columnspan=2, pady=(5, 0))
    
    def create_current_scan_section(self):
        """Erstellt den Bereich für den aktuellen Scan"""
        # Current-Scan-Frame
        self.current_frame = ttk.LabelFrame(self.main_frame, text="✅ AKTUELLER SCAN", padding="10")
        self.current_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        self.current_frame.columnconfigure(1, weight=1)
        
        # Datenfelder
        self.create_data_labels()
        
        # Eingabefelder
        self.create_input_fields()
        
        # Verstecke initial
        self.current_frame.grid_remove()
    
    def create_data_labels(self):
        """Erstellt die Datenanzeigefelder"""
        row = 0
        
        # Charge
        ttk.Label(self.current_frame, text="Charge:", font=("Arial", 11, "bold")).grid(row=row, column=0, sticky=tk.W, pady=2)
        self.charge_label = ttk.Label(self.current_frame, text="", font=("Arial", 11))
        self.charge_label.grid(row=row, column=1, sticky=tk.W, padx=(10, 0), pady=2)
        row += 1
        
        # Material
        ttk.Label(self.current_frame, text="Material:", font=("Arial", 11, "bold")).grid(row=row, column=0, sticky=tk.W, pady=2)
        self.material_label = ttk.Label(self.current_frame, text="", font=("Arial", 11))
        self.material_label.grid(row=row, column=1, sticky=tk.W, padx=(10, 0), pady=2)
        row += 1
        
        # Materialkurztext
        ttk.Label(self.current_frame, text="Materialkurztext:", font=("Arial", 11, "bold")).grid(row=row, column=0, sticky=tk.W, pady=2)
        self.kurztext_label = ttk.Label(self.current_frame, text="", font=("Arial", 11))
        self.kurztext_label.grid(row=row, column=1, sticky=tk.W, padx=(10, 0), pady=2)
        row += 1
        
        # Länge
        ttk.Label(self.current_frame, text="Länge:", font=("Arial", 11, "bold")).grid(row=row, column=0, sticky=tk.W, pady=2)
        self.laenge_label = ttk.Label(self.current_frame, text="", font=("Arial", 11))
        self.laenge_label.grid(row=row, column=1, sticky=tk.W, padx=(10, 0), pady=2)
        row += 1
        
        # Breite
        ttk.Label(self.current_frame, text="Breite:", font=("Arial", 11, "bold")).grid(row=row, column=0, sticky=tk.W, pady=2)
        self.breite_label = ttk.Label(self.current_frame, text="", font=("Arial", 11))
        self.breite_label.grid(row=row, column=1, sticky=tk.W, padx=(10, 0), pady=2)
        row += 1
        
        # Fläche
        ttk.Label(self.current_frame, text="Fläche:", font=("Arial", 11, "bold")).grid(row=row, column=0, sticky=tk.W, pady=2)
        self.flaeche_label = ttk.Label(self.current_frame, text="", font=("Arial", 11))
        self.flaeche_label.grid(row=row, column=1, sticky=tk.W, padx=(10, 0), pady=2)
        row += 1
        
        # Trennlinie
        separator = ttk.Separator(self.current_frame, orient='horizontal')
        separator.grid(row=row, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)
        row += 1
        
        self.input_start_row = row
    
    def create_input_fields(self):
        """Erstellt die Eingabefelder - werden dynamisch je nach Typ angezeigt"""
        # Initialisiere alle Variablen
        self.fach_var = tk.StringVar()
        self.bemerkung_var = tk.StringVar()
        self.breite_kontrolliert_var = tk.StringVar()
        self.zahlmenge_var = tk.StringVar()
        
        # Container für dynamische Eingabefelder
        self.input_container = ttk.Frame(self.current_frame)
        self.input_container.grid(row=self.input_start_row, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)
        self.input_container.columnconfigure(1, weight=1)
        
        # Widgets werden dynamisch erstellt
        self.input_widgets = {}
    
    def create_rolle_inputs(self):
        """Erstellt Eingabefelder für Rollen"""
        self.clear_input_widgets()
        row = 0
        
        # Fach-Eingabe (PFLICHT)
        fach_label = ttk.Label(self.input_container, text="🏷️ Fach (Lagerort):", font=("Arial", 11, "bold"))
        fach_label.grid(row=row, column=0, sticky=tk.W, pady=5)
        fach_entry = ttk.Entry(self.input_container, 
                              textvariable=self.fach_var,
                              font=("Arial", 12),
                              width=20)
        fach_entry.grid(row=row, column=1, sticky=tk.W, padx=(10, 0), pady=5)
        fach_entry.bind('<Return>', self.save_current_scan)
        fach_entry.bind('<KeyRelease>', self.on_field_change)
        self.input_widgets['fach_label'] = fach_label
        self.input_widgets['fach_entry'] = fach_entry
        row += 1
        
        # Breite kontrolliert (PFLICHT)
        breite_label = ttk.Label(self.input_container, text="📏 Breite kontrolliert (mm):", font=("Arial", 11, "bold"))
        breite_label.grid(row=row, column=0, sticky=tk.W, pady=5)
        breite_entry = ttk.Entry(self.input_container, 
                                textvariable=self.breite_kontrolliert_var,
                                font=("Arial", 12),
                                width=20)
        breite_entry.grid(row=row, column=1, sticky=tk.W, padx=(10, 0), pady=5)
        breite_entry.bind('<Return>', self.save_current_scan)
        breite_entry.bind('<KeyRelease>', self.on_field_change)
        self.input_widgets['breite_label'] = breite_label
        self.input_widgets['breite_entry'] = breite_entry
        row += 1
        
        # Bemerkung (optional)
        bemerkung_label = ttk.Label(self.input_container, text="📝 Bemerkung (optional):", font=("Arial", 11, "bold"))
        bemerkung_label.grid(row=row, column=0, sticky=tk.W, pady=5)
        bemerkung_entry = ttk.Entry(self.input_container, 
                                   textvariable=self.bemerkung_var,
                                   font=("Arial", 12),
                                   width=50)
        bemerkung_entry.grid(row=row, column=1, sticky=tk.W, padx=(10, 0), pady=5)
        bemerkung_entry.bind('<Return>', self.save_current_scan)
        bemerkung_entry.bind('<KeyRelease>', self.on_field_change)
        self.input_widgets['bemerkung_label'] = bemerkung_label
        self.input_widgets['bemerkung_entry'] = bemerkung_entry
        row += 1
        
        # Speichern-Button
        save_button = ttk.Button(self.input_container, text="💾 Speichern", command=self.save_current_scan)
        save_button.grid(row=row, column=1, sticky=tk.W, padx=(10, 0), pady=10)
        self.input_widgets['save_button'] = save_button
        
        # Fokus auf Fach-Eingabe
        fach_entry.focus_set()
    
    def create_granulat_inputs(self):
        """Erstellt Eingabefelder für Granulat"""
        self.clear_input_widgets()
        row = 0
        
        # Zählmenge (PFLICHT)
        zahlmenge_label = ttk.Label(self.input_container, text="⚖️ Zählmenge (KG):", font=("Arial", 11, "bold"))
        zahlmenge_label.grid(row=row, column=0, sticky=tk.W, pady=5)
        zahlmenge_entry = ttk.Entry(self.input_container, 
                                   textvariable=self.zahlmenge_var,
                                   font=("Arial", 12),
                                   width=20)
        zahlmenge_entry.grid(row=row, column=1, sticky=tk.W, padx=(10, 0), pady=5)
        zahlmenge_entry.bind('<Return>', self.save_current_scan)
        zahlmenge_entry.bind('<KeyRelease>', self.on_field_change)
        self.input_widgets['zahlmenge_label'] = zahlmenge_label
        self.input_widgets['zahlmenge_entry'] = zahlmenge_entry
        row += 1
        
        # Bemerkung (optional)
        bemerkung_label = ttk.Label(self.input_container, text="📝 Bemerkung (optional):", font=("Arial", 11, "bold"))
        bemerkung_label.grid(row=row, column=0, sticky=tk.W, pady=5)
        bemerkung_entry = ttk.Entry(self.input_container, 
                                   textvariable=self.bemerkung_var,
                                   font=("Arial", 12),
                                   width=50)
        bemerkung_entry.grid(row=row, column=1, sticky=tk.W, padx=(10, 0), pady=5)
        bemerkung_entry.bind('<Return>', self.save_current_scan)
        bemerkung_entry.bind('<KeyRelease>', self.on_field_change)
        self.input_widgets['bemerkung_label'] = bemerkung_label
        self.input_widgets['bemerkung_entry'] = bemerkung_entry
        row += 1
        
        # Speichern-Button
        save_button = ttk.Button(self.input_container, text="💾 Speichern", command=self.save_current_scan)
        save_button.grid(row=row, column=1, sticky=tk.W, padx=(10, 0), pady=10)
        self.input_widgets['save_button'] = save_button
        
        # Fokus auf Zählmenge-Eingabe
        zahlmenge_entry.focus_set()
    
    def clear_input_widgets(self):
        """Löscht alle dynamischen Eingabe-Widgets"""
        for widget in self.input_widgets.values():
            widget.destroy()
        self.input_widgets.clear()
    
    def validiere_breite(self, breite_text):
        """Validiert Breite-Eingabe (1-4 stellige Zahl)"""
        try:
            breite = int(breite_text.strip())
            if 1 <= breite <= 9999:  # 1-4 stellig
                return True, breite
            else:
                return False, "Breite muss zwischen 1 und 9999 mm liegen!"
        except ValueError:
            return False, "Breite muss eine gültige Zahl sein!"
    
    def validiere_gewicht(self, gewicht_text):
        """Validiert Gewichts-Eingabe (positive Dezimalzahl)"""
        try:
            gewicht = float(gewicht_text.strip().replace(',', '.'))
            if gewicht > 0:
                return True, gewicht
            else:
                return False, "Zählmenge muss größer als 0 sein!"
        except ValueError:
            return False, "Zählmenge muss eine gültige Zahl sein!"
    
    def create_list_section(self):
        """Erstellt die Liste der gescannten Artikel"""
        # List-Frame
        list_frame = ttk.LabelFrame(self.main_frame, text="📋 GESCANNTE ARTIKEL", padding="10")
        list_frame.grid(row=3, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(1, weight=1)
        self.main_frame.rowconfigure(3, weight=1)
        
        # Artikel-Anzahl (ohne Suchbereich)
        count_frame = ttk.Frame(list_frame)
        count_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.count_label = ttk.Label(count_frame, text="0 Artikel", font=("Arial", 12, "bold"))
        self.count_label.pack(side=tk.LEFT)
        
        # Treeview für Artikelliste
        columns = ('Zeit', 'Charge', 'Material', 'Typ', 'Fach', 'Status')
        self.tree = ttk.Treeview(list_frame, columns=columns, show='headings', height=15)
        
        # Spalten konfigurieren
        self.tree.heading('Zeit', text='Zeit')
        self.tree.heading('Charge', text='Charge')
        self.tree.heading('Material', text='Material')
        self.tree.heading('Typ', text='Typ')
        self.tree.heading('Fach', text='Fach')
        self.tree.heading('Status', text='Status')
        
        # Spaltenbreiten
        self.tree.column('Zeit', width=80, minwidth=70)
        self.tree.column('Charge', width=100, minwidth=80)
        self.tree.column('Material', width=100, minwidth=80)
        self.tree.column('Typ', width=80, minwidth=70)
        self.tree.column('Fach', width=80, minwidth=60)
        self.tree.column('Status', width=120, minwidth=100)
        
        self.tree.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.tree.yview)
        v_scrollbar.grid(row=1, column=1, sticky=(tk.N, tk.S))
        self.tree.configure(yscrollcommand=v_scrollbar.set)
        
        h_scrollbar = ttk.Scrollbar(list_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        h_scrollbar.grid(row=2, column=0, sticky=(tk.W, tk.E))
        self.tree.configure(xscrollcommand=h_scrollbar.set)
        
        # Kontextmenü
        self.tree.bind('<Button-3>', self.show_context_menu)
        self.tree.bind('<Double-1>', self.edit_entry)
    
    def create_button_section(self):
        """Erstellt die Button-Leiste"""
        button_frame = ttk.Frame(self.main_frame)
        button_frame.grid(row=4, column=0, sticky=(tk.W, tk.E), pady=(10, 0))
        
        # Export-Button
        export_button = ttk.Button(button_frame, text="💾 Inventur exportieren", command=self.export_inventur)
        export_button.grid(row=0, column=0, padx=(0, 10))
        
        # Vollbild-Toggle
        fullscreen_button = ttk.Button(button_frame, text="🖥️ Vollbild", command=self.toggle_fullscreen)
        fullscreen_button.grid(row=0, column=1, padx=(0, 10))
        
        # Beenden-Button
        exit_button = ttk.Button(button_frame, text="❌ Programm beenden", command=self.quit_app)
        exit_button.grid(row=0, column=2)
    
    def create_status_bar(self):
        """Erstellt die Status-Leiste"""
        self.status_var = tk.StringVar()
        self.status_var.set("Bereit zum Scannen...")
        
        status_bar = ttk.Label(self.root, textvariable=self.status_var, 
                              relief=tk.SUNKEN, anchor=tk.W, font=("Arial", 9))
        status_bar.grid(row=1, column=0, sticky=(tk.W, tk.E))
    
    def bind_shortcuts(self):
        """Bindet Tastenkürzel"""
        self.root.bind('<Control-z>', lambda e: self.undo_last_action())
        self.root.bind('<Control-Z>', lambda e: self.undo_last_action())
        self.root.bind('<Control-s>', lambda e: self.manual_save())
        self.root.bind('<Control-S>', lambda e: self.manual_save())
        self.root.bind('<Escape>', lambda e: self.reset_scan())
        self.root.bind('<F11>', lambda e: self.toggle_fullscreen())
        
        # Fokus immer zurück zum Scan-Feld
        self.root.bind('<FocusIn>', self.ensure_scan_focus)
    
    def suche_charge(self, charge_nummer):
        """Sucht Charge in beiden Tabellenblättern und gibt Typ zurück"""
        # 1. Zuerst in Rollen suchen
        if self.df_rollen is not None:
            ergebnis_rolle = self.df_rollen[self.df_rollen['Charge'] == str(charge_nummer)]
            if not ergebnis_rolle.empty:
                return ('ROLLE', ergebnis_rolle.iloc[0].to_dict())
        
        # 2. Dann in Granulate suchen
        if self.df_granulate is not None:
            ergebnis_granulat = self.df_granulate[self.df_granulate['Charge'] == str(charge_nummer)]
            if not ergebnis_granulat.empty:
                return ('GRANULAT', ergebnis_granulat.iloc[0].to_dict())
        
        # 3. Nicht gefunden
        return ('NICHT_GEFUNDEN', None)
    
    def process_scan(self):
        """Verarbeitet einen gescannten Barcode"""
        charge = self.scan_var.get().strip()
        
        if not charge:
            self.status_var.set("Bitte Barcode eingeben oder scannen")
            return
        
        # Prüfe ob Charge bereits gescannt wurde
        if self.is_already_scanned(charge):
            messagebox.showwarning("Bereits gescannt", 
                f"Die Ware mit Charge {charge} wurde bereits eingescannt!\n\n"
                f"Bitte prüfen Sie die Liste der gescannten Artikel.")
            self.reset_scan()
            return
        
        self.logger.info(f"Scan verarbeitet: {charge}")
        
        # Suche mit neuer Typ-Erkennung
        try:
            # Suche nach Charge als String (behält führende Nullen)
            typ, data = self.suche_charge(charge)
            
            # Falls nicht gefunden, versuche auch ohne führende Nullen
            if typ == 'NICHT_GEFUNDEN':
                try:
                    charge_int = str(int(charge))  # Entfernt führende Nullen
                    typ, data = self.suche_charge(charge_int)
                    if typ != 'NICHT_GEFUNDEN':
                        charge = charge_int  # Verwende bereinigte Charge
                except ValueError:
                    pass
            
            if typ == 'ROLLE':
                # Rolle gefunden
                self.show_found_rolle(data, charge)
            elif typ == 'GRANULAT':
                # Granulat gefunden
                self.show_found_granulat(data, charge)
            else:
                # Ware nicht gefunden
                self.show_not_found_dialog(charge)
                
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler bei der Suche: {e}")
            self.logger.error(f"Fehler bei Suche: {e}")
            self.reset_scan()
    
    def is_already_scanned(self, charge):
        """Prüft ob eine Charge bereits gescannt wurde"""
        # Prüfe in allen Listen
        all_lists = [
            self.inventur_rollen_data,
            self.inventur_granulat_data,
            self.nicht_gefunden_rollen_data,
            self.nicht_gefunden_granulat_data
        ]
        
        for item_list in all_lists:
            for item in item_list:
                if str(item.get('charge', '')) == str(charge):
                    return True
        
        return False
    
    def show_found_rolle(self, item, charge):
        """Zeigt gefundene Rolle an (BLAU)"""
        self.current_type = 'ROLLE'
        self.current_scan = {
            'charge': charge,
            'material': str(item.get('Material', '')),
            'kurztext': str(item.get('Materialkurztext', '')),
            'laenge': float(item.get('Länge m', 0)),
            'breite_original': int(item.get('Breite mm', 0)),  # Original aus Arbeitstabelle
            'flaeche': float(item.get('Frei verwendbar', 0)),
            'fach_original': str(item.get('Fach', '') if pd.notna(item.get('Fach')) else ''),  # Original aus Arbeitstabelle
            'status': 'gefunden',
            'typ': 'ROLLE'
        }
        
        # Ändere Hintergrundfarbe zu BLAU
        self.current_frame.config(text="🔵 ROLLE GESCANNT")
        
        # Aktualisiere Labels (OHNE Breite mm Anzeige!)
        self.charge_label.config(text=charge)
        self.material_label.config(text=self.current_scan['material'])
        self.kurztext_label.config(text=self.current_scan['kurztext'])
        self.laenge_label.config(text=f"{self.current_scan['laenge']:.2f} m")
        self.breite_label.config(text="")  # NICHT anzeigen!
        self.flaeche_label.config(text=f"{self.current_scan['flaeche']:.2f} m²")
        
        # Zeige Current-Scan-Frame mit blauem Hintergrund
        self.current_frame.grid()
        
        # Erstelle Rollen-Eingabefelder
        self.create_rolle_inputs()
        
        # Status aktualisieren
        self.status_var.set(f"🔵 Rolle gefunden: {self.current_scan['kurztext']}")
        
        # Scan-Feld leeren
        self.scan_var.set("")
    
    def show_found_granulat(self, item, charge):
        """Zeigt gefundenes Granulat an (GELB)"""
        self.current_type = 'GRANULAT'
        self.current_scan = {
            'charge': charge,
            'material': str(item.get('Material', '')),
            'kurztext': str(item.get('Materialkurztext', '')),
            'frei_verwendbar_kg': float(item.get('Frei verwendbar', 0)),  # Soll-Gewicht
            'status': 'gefunden',
            'typ': 'GRANULAT'
        }
        
        # Ändere Hintergrundfarbe zu GELB
        self.current_frame.config(text="🟨 GRANULAT GESCANNT")
        
        # Aktualisiere Labels (nur relevante für Granulat)
        self.charge_label.config(text=charge)
        self.material_label.config(text=self.current_scan['material'])
        self.kurztext_label.config(text=self.current_scan['kurztext'])
        self.laenge_label.config(text="")  # Nicht relevant für Granulat
        self.breite_label.config(text="")  # Nicht relevant für Granulat
        self.flaeche_label.config(text=f"{self.current_scan['frei_verwendbar_kg']:.2f} KG")
        
        # Zeige Current-Scan-Frame mit gelbem Hintergrund
        self.current_frame.grid()
        
        # Erstelle Granulat-Eingabefelder
        self.create_granulat_inputs()
        
        # Status aktualisieren
        self.status_var.set(f"🟨 Granulat gefunden: {self.current_scan['kurztext']}")
        
        # Scan-Feld leeren
        self.scan_var.set("")
    
    def show_not_found_dialog(self, charge):
        """Zeigt Dialog für nicht gefundene Ware (V2 mit Typ-Auswahl)"""
        dialog = NotFoundDialog(self.root, charge)
        
        if dialog.result:
            # Setze aktuellen Scan mit manuellen Daten
            self.current_scan = dialog.result.copy()
            self.current_type = dialog.result['typ']
            
            # Füge Zeitstempel hinzu
            self.current_scan['zeitstempel'] = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
            
            # Speichere in entsprechende Liste basierend auf Typ
            if self.current_type == 'ROLLE':
                self.nicht_gefunden_rollen_data.append(self.current_scan.copy())
                typ_icon = "🔵"
            elif self.current_type == 'GRANULAT':
                self.nicht_gefunden_granulat_data.append(self.current_scan.copy())
                typ_icon = "🟨"
            
            # Zur Undo-Liste hinzufügen
            self.undo_stack.append(('add', self.current_scan.copy(), self.current_type))
            if len(self.undo_stack) > 50:
                self.undo_stack.pop(0)
            
            # In Excel speichern (Auto-Save)
            if self.config.get('auto_save', True):
                self.save_to_excel()
            
            # Liste aktualisieren
            self.update_list()
            
            # Status aktualisieren
            total_rollen = len(self.inventur_rollen_data) + len(self.nicht_gefunden_rollen_data)
            total_granulat = len(self.inventur_granulat_data) + len(self.nicht_gefunden_granulat_data)
            total_scans = total_rollen + total_granulat
            
            self.status_var.set(f"{typ_icon} Nicht gefundene Ware gespeichert. Gesamt: {total_scans} ({total_rollen} Rollen, {total_granulat} Granulate)")
            self.logger.info(f"{self.current_type} nicht gefunden gespeichert: {self.current_scan['charge']}")
            
            # Reset für nächsten Scan
            self.reset_scan()
        else:
            # Abgebrochen
            self.reset_scan()
    
    def save_current_scan(self, event=None):
        """Speichert den aktuellen Scan (Rolle oder Granulat)"""
        if not self.current_scan or not self.current_type:
            return
        
        if self.current_type == 'ROLLE':
            # Validierung für Rollen
            fach = self.fach_var.get().strip()
            if not fach:
                messagebox.showwarning("Warnung", "⚠️ Fach (Lagerort) ist ein Pflichtfeld für Rollen!")
                if 'fach_entry' in self.input_widgets:
                    self.input_widgets['fach_entry'].focus_set()
                return
            
            breite_text = self.breite_kontrolliert_var.get().strip()
            if not breite_text:
                messagebox.showwarning("Warnung", "⚠️ Breite kontrolliert ist ein Pflichtfeld!")
                if 'breite_entry' in self.input_widgets:
                    self.input_widgets['breite_entry'].focus_set()
                return
            
            # Validiere Breite
            valid, result = self.validiere_breite(breite_text)
            if not valid:
                messagebox.showerror("Fehler", f"⚠️ {result}")
                if 'breite_entry' in self.input_widgets:
                    self.input_widgets['breite_entry'].focus_set()
                return
            
            # Füge Rollen-spezifische Daten hinzu
            self.current_scan['fach_kontrolliert'] = fach
            self.current_scan['breite_kontrolliert'] = result
            self.current_scan['bemerkung'] = self.bemerkung_var.get().strip()
            
        elif self.current_type == 'GRANULAT':
            # Validierung für Granulat
            zahlmenge_text = self.zahlmenge_var.get().strip()
            if not zahlmenge_text:
                messagebox.showwarning("Warnung", "⚠️ Zählmenge ist ein Pflichtfeld!")
                if 'zahlmenge_entry' in self.input_widgets:
                    self.input_widgets['zahlmenge_entry'].focus_set()
                return
            
            # Validiere Gewicht
            valid, result = self.validiere_gewicht(zahlmenge_text)
            if not valid:
                messagebox.showerror("Fehler", f"⚠️ {result}")
                if 'zahlmenge_entry' in self.input_widgets:
                    self.input_widgets['zahlmenge_entry'].focus_set()
                return
            
            # Füge Granulat-spezifische Daten hinzu
            self.current_scan['zahlmenge_kg'] = result
            self.current_scan['bemerkung'] = self.bemerkung_var.get().strip()
        
        # Speichere in Datenstrukturen
        self.save_scan_to_data()
        
        # Reset für nächsten Scan
        self.reset_scan()
    
    def save_scan_to_data(self):
        """Speichert Scan in Datenstrukturen und Excel"""
        if not self.current_scan or not self.current_type:
            return
        
        # Zeitstempel hinzufügen
        self.current_scan['zeitstempel'] = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
        
        # Zu entsprechender Liste hinzufügen basierend auf Typ und Status
        if self.current_type == 'ROLLE':
            if self.current_scan['status'] == 'gefunden':
                self.inventur_rollen_data.append(self.current_scan.copy())
            else:
                self.nicht_gefunden_rollen_data.append(self.current_scan.copy())
        elif self.current_type == 'GRANULAT':
            if self.current_scan['status'] == 'gefunden':
                self.inventur_granulat_data.append(self.current_scan.copy())
            else:
                self.nicht_gefunden_granulat_data.append(self.current_scan.copy())
        
        # Zur Undo-Liste hinzufügen
        self.undo_stack.append(('add', self.current_scan.copy(), self.current_type))
        if len(self.undo_stack) > 50:  # Begrenze Undo-Stack
            self.undo_stack.pop(0)
        
        # In Excel speichern (Auto-Save)
        if self.config.get('auto_save', True):
            self.save_to_excel()
        
        # Liste aktualisieren
        self.update_list()
        
        # Status aktualisieren
        total_rollen = len(self.inventur_rollen_data) + len(self.nicht_gefunden_rollen_data)
        total_granulat = len(self.inventur_granulat_data) + len(self.nicht_gefunden_granulat_data)
        total_scans = total_rollen + total_granulat
        
        typ_icon = "🔵" if self.current_type == 'ROLLE' else "🟨"
        self.status_var.set(f"{typ_icon} Artikel gespeichert. Gesamt: {total_scans} ({total_rollen} Rollen, {total_granulat} Granulate)")
        
        self.logger.info(f"{self.current_type} gespeichert: {self.current_scan['charge']}")
    
    def reset_scan(self):
        """Setzt den aktuellen Scan zurück"""
        self.current_scan = None
        self.current_type = None
        self.scan_var.set("")
        self.fach_var.set("")
        self.bemerkung_var.set("")
        self.breite_kontrolliert_var.set("")
        self.zahlmenge_var.set("")
        
        # Lösche dynamische Eingabefelder
        self.clear_input_widgets()
        
        # Verstecke Current-Scan-Frame
        self.current_frame.grid_remove()
        
        # Fokus zurück zu Scan-Eingabe
        self.scan_entry.focus_set()
        
        self.status_var.set("Bereit zum Scannen...")
    
    def on_field_change(self, event=None):
        """Wird aufgerufen wenn Fach oder Bemerkung geändert wird"""
        if self.config.get('auto_save', True) and self.current_scan:
            # Auto-Save bei Feldeingabe (mit kleiner Verzögerung)
            self.root.after(1000, self.delayed_save)
    
    def delayed_save(self):
        """Verzögertes Speichern für Auto-Save"""
        if self.current_scan and self.fach_var.get().strip():
            # Nur speichern wenn Fach ausgefüllt ist
            pass  # Wird durch save_current_scan() erledigt
    
    def update_list(self):
        """Aktualisiert die Artikelliste mit neuer Struktur"""
        # Lösche alle Einträge
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Kombiniere alle Listen mit Typ-Information
        all_items = []
        
        # Rollen - gefunden
        for item in self.inventur_rollen_data:
            all_items.append((item, 'Gefunden', 'ROLLE'))
        
        # Rollen - nicht gefunden
        for item in self.nicht_gefunden_rollen_data:
            all_items.append((item, 'Nicht gefunden', 'ROLLE'))
        
        # Granulat - gefunden
        for item in self.inventur_granulat_data:
            all_items.append((item, 'Gefunden', 'GRANULAT'))
        
        # Granulat - nicht gefunden
        for item in self.nicht_gefunden_granulat_data:
            all_items.append((item, 'Nicht gefunden', 'GRANULAT'))
        
        # Sortiere nach Zeitstempel (neueste zuerst)
        all_items.sort(key=lambda x: x[0]['zeitstempel'], reverse=True)
        
        # Füge zur TreeView hinzu
        for item_data, status, typ in all_items:
            # Typ-Icon
            typ_icon = "🔵 Rolle" if typ == 'ROLLE' else "🟨 Granu"
            
            # Fach-Information (unterschiedlich je nach Typ)
            if typ == 'ROLLE':
                fach_info = item_data.get('fach_kontrolliert', item_data.get('fach', ''))
            else:
                fach_info = '-'  # Granulat hat kein Fach
            
            values = (
                item_data['zeitstempel'].split()[1],  # Nur Zeit anzeigen
                item_data['charge'],
                item_data['material'],
                typ_icon,
                fach_info,
                status
            )
            
            item_id = self.tree.insert('', 'end', values=values)
            
            # Status-Icon setzen
            if status == 'Nicht gefunden':
                self.tree.set(item_id, 'Status', '⚠️ Nicht gefunden')
            else:
                self.tree.set(item_id, 'Status', '✅ Gefunden')
        
        # Anzahl aktualisieren
        total_rollen = len(self.inventur_rollen_data) + len(self.nicht_gefunden_rollen_data)
        total_granulat = len(self.inventur_granulat_data) + len(self.nicht_gefunden_granulat_data)
        total = total_rollen + total_granulat
        
        self.count_label.config(text=f"{total} Artikel (🔵 {total_rollen} Rollen, 🟨 {total_granulat} Granulate)")
    
    def save_to_excel(self):
        """Speichert Daten in separate Excel-Dateien für Rollen und Granulat"""
        try:
            # Speichere Rollen-Datei
            self.save_rollen_excel()
            
            # Speichere Granulat-Datei
            self.save_granulat_excel()
            
            self.logger.info("Excel-Dateien gespeichert")
            
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Speichern der Excel-Dateien:\n{e}")
            self.logger.error(f"Fehler beim Speichern: {e}")
    
    def save_rollen_excel(self):
        """Speichert Rollen-Daten in Inventur_Rollen.xlsx"""
        # Erstelle oder lade Workbook
        if self.inventur_rollen_path.exists():
            wb = load_workbook(self.inventur_rollen_path)
        else:
            wb = Workbook()
            # Entferne Standard-Sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
        
        # Erstelle/aktualisiere Inventur-Sheet
        if 'Inventur' in wb.sheetnames:
            wb.remove(wb['Inventur'])
        
        ws_inventur = wb.create_sheet('Inventur')
        
        # Header für Rollen-Inventur (erweiterte Struktur)
        headers = ['Datum/Uhrzeit', 'Charge', 'Material', 'Materialkurztext', 
                  'Länge m', 'Fläche m²', 'Breite mm', 'Breite kontrolliert', 
                  'Fach', 'Fach kontrolliert', 'Bemerkung']
        ws_inventur.append(headers)
        
        # Daten für Rollen-Inventur
        for item in self.inventur_rollen_data:
            bemerkung = item.get('bemerkung', '')
            if bemerkung == 'nan' or str(bemerkung).lower() == 'nan':
                bemerkung = ''
            
            row = [
                item['zeitstempel'],
                item['charge'],
                item['material'],
                item['kurztext'],
                item['laenge'],
                item['flaeche'],
                item.get('breite_original', ''),  # Original aus Arbeitstabelle
                item.get('breite_kontrolliert', ''),  # Vom Nutzer eingegeben
                item.get('fach_original', ''),  # Original aus Arbeitstabelle
                item.get('fach_kontrolliert', ''),  # Vom Nutzer eingegeben
                bemerkung
            ]
            ws_inventur.append(row)
        
        # Erstelle/aktualisiere Nicht_gefunden-Sheet
        if 'Nicht_gefunden' in wb.sheetnames:
            wb.remove(wb['Nicht_gefunden'])
        
        ws_nicht_gefunden = wb.create_sheet('Nicht_gefunden')
        ws_nicht_gefunden.append(headers)
        
        # Daten für Nicht_gefunden Rollen
        for item in self.nicht_gefunden_rollen_data:
            bemerkung = item.get('bemerkung', '')
            if bemerkung == 'nan' or str(bemerkung).lower() == 'nan':
                bemerkung = ''
            
            row = [
                item['zeitstempel'],
                item['charge'],
                item['material'],
                item['kurztext'],
                item.get('laenge', ''),
                item.get('flaeche', ''),
                item.get('breite_original', ''),
                item.get('breite_kontrolliert', ''),
                item.get('fach_original', ''),
                item.get('fach_kontrolliert', ''),
                bemerkung
            ]
            ws_nicht_gefunden.append(row)
        
        # Formatiere Charge-Spalte als Text
        from openpyxl.utils import get_column_letter
        charge_col = get_column_letter(2)  # Spalte B
        
        # Inventur-Sheet
        for row in range(2, len(self.inventur_rollen_data) + 2):
            cell = ws_inventur[f'{charge_col}{row}']
            cell.number_format = '@'
        
        # Nicht_gefunden-Sheet
        for row in range(2, len(self.nicht_gefunden_rollen_data) + 2):
            cell = ws_nicht_gefunden[f'{charge_col}{row}']
            cell.number_format = '@'
        
        # Speichern
        wb.save(self.inventur_rollen_path)
    
    def save_granulat_excel(self):
        """Speichert Granulat-Daten in Inventur_Granulat.xlsx"""
        # Erstelle oder lade Workbook
        if self.inventur_granulat_path.exists():
            wb = load_workbook(self.inventur_granulat_path)
        else:
            wb = Workbook()
            # Entferne Standard-Sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
        
        # Erstelle/aktualisiere Inventur-Sheet
        if 'Inventur' in wb.sheetnames:
            wb.remove(wb['Inventur'])
        
        ws_inventur = wb.create_sheet('Inventur')
        
        # Header für Granulat-Inventur
        headers = ['Datum/Uhrzeit', 'Charge', 'Material', 'Materialkurztext', 
                  'Frei verwendbar (KG)', 'Zählmenge (KG)', 'Bemerkung']
        ws_inventur.append(headers)
        
        # Daten für Granulat-Inventur
        for item in self.inventur_granulat_data:
            bemerkung = item.get('bemerkung', '')
            if bemerkung == 'nan' or str(bemerkung).lower() == 'nan':
                bemerkung = ''
            
            row = [
                item['zeitstempel'],
                item['charge'],
                item['material'],
                item['kurztext'],
                item.get('frei_verwendbar_kg', ''),  # Soll-Gewicht
                item.get('zahlmenge_kg', ''),  # Ist-Gewicht
                bemerkung
            ]
            ws_inventur.append(row)
        
        # Erstelle/aktualisiere Nicht_gefunden-Sheet
        if 'Nicht_gefunden' in wb.sheetnames:
            wb.remove(wb['Nicht_gefunden'])
        
        ws_nicht_gefunden = wb.create_sheet('Nicht_gefunden')
        ws_nicht_gefunden.append(headers)
        
        # Daten für Nicht_gefunden Granulat
        for item in self.nicht_gefunden_granulat_data:
            bemerkung = item.get('bemerkung', '')
            if bemerkung == 'nan' or str(bemerkung).lower() == 'nan':
                bemerkung = ''
            
            row = [
                item['zeitstempel'],
                item['charge'],
                item['material'],
                item['kurztext'],
                item.get('frei_verwendbar_kg', ''),
                item.get('zahlmenge_kg', ''),
                bemerkung
            ]
            ws_nicht_gefunden.append(row)
        
        # Formatiere Charge-Spalte als Text
        from openpyxl.utils import get_column_letter
        charge_col = get_column_letter(2)  # Spalte B
        
        # Inventur-Sheet
        for row in range(2, len(self.inventur_granulat_data) + 2):
            cell = ws_inventur[f'{charge_col}{row}']
            cell.number_format = '@'
        
        # Nicht_gefunden-Sheet
        for row in range(2, len(self.nicht_gefunden_granulat_data) + 2):
            cell = ws_nicht_gefunden[f'{charge_col}{row}']
            cell.number_format = '@'
        
        # Speichern
        wb.save(self.inventur_granulat_path)
    
    def load_existing_inventur(self):
        """Lädt bestehende Inventur-Daten aus beiden V2-Dateien"""
        total_loaded = 0
        
        # Lade Rollen-Inventur
        total_loaded += self.load_existing_rollen()
        
        # Lade Granulat-Inventur
        total_loaded += self.load_existing_granulat()
        
        if total_loaded > 0:
            # Liste aktualisieren
            self.update_list()
            
            total_rollen = len(self.inventur_rollen_data) + len(self.nicht_gefunden_rollen_data)
            total_granulat = len(self.inventur_granulat_data) + len(self.nicht_gefunden_granulat_data)
            
            self.status_var.set(f"Bestehende Inventur geladen: {total_rollen} Rollen, {total_granulat} Granulate")
            self.logger.info(f"Bestehende Inventur geladen: {total_rollen} Rollen, {total_granulat} Granulate")
    
    def load_existing_rollen(self):
        """Lädt bestehende Rollen-Inventur"""
        if not self.inventur_rollen_path.exists():
            return 0
        
        loaded_count = 0
        try:
            # Lade Rollen Inventur-Sheet
            df_inventur = pd.read_excel(self.inventur_rollen_path, sheet_name='Inventur', dtype={'Charge': str})
            for _, row in df_inventur.iterrows():
                bemerkung = row.get('Bemerkung', '')
                if pd.isna(bemerkung) or str(bemerkung).lower() == 'nan':
                    bemerkung = ''
                
                item = {
                    'zeitstempel': str(row.get('Datum/Uhrzeit', '')),
                    'charge': str(row.get('Charge', '')),
                    'material': str(row.get('Material', '')),
                    'kurztext': str(row.get('Materialkurztext', '')),
                    'laenge': float(row.get('Länge m', 0)),
                    'flaeche': float(row.get('Fläche m²', 0)),
                    'breite_original': int(row.get('Breite mm', 0)),
                    'breite_kontrolliert': int(row.get('Breite kontrolliert', 0)),
                    'fach_original': str(row.get('Fach', '')),
                    'fach_kontrolliert': str(row.get('Fach kontrolliert', '')),
                    'bemerkung': str(bemerkung),
                    'status': 'gefunden',
                    'typ': 'ROLLE'
                }
                self.inventur_rollen_data.append(item)
                loaded_count += 1
            
            # Lade Rollen Nicht_gefunden-Sheet
            try:
                df_nicht_gefunden = pd.read_excel(self.inventur_rollen_path, sheet_name='Nicht_gefunden', dtype={'Charge': str})
                for _, row in df_nicht_gefunden.iterrows():
                    bemerkung = row.get('Bemerkung', '')
                    if pd.isna(bemerkung) or str(bemerkung).lower() == 'nan':
                        bemerkung = ''
                    
                    item = {
                        'zeitstempel': str(row.get('Datum/Uhrzeit', '')),
                        'charge': str(row.get('Charge', '')),
                        'material': str(row.get('Material', '')),
                        'kurztext': str(row.get('Materialkurztext', '')),
                        'laenge': float(row.get('Länge m', 0)),
                        'flaeche': float(row.get('Fläche m²', 0)),
                        'breite_original': int(row.get('Breite mm', 0)),
                        'breite_kontrolliert': int(row.get('Breite kontrolliert', 0)),
                        'fach_original': str(row.get('Fach', '')),
                        'fach_kontrolliert': str(row.get('Fach kontrolliert', '')),
                        'bemerkung': str(bemerkung),
                        'status': 'nicht_gefunden',
                        'typ': 'ROLLE'
                    }
                    self.nicht_gefunden_rollen_data.append(item)
                    loaded_count += 1
            except:
                pass  # Sheet existiert noch nicht
                
        except Exception as e:
            self.logger.error(f"Fehler beim Laden der Rollen-Inventur: {e}")
        
        return loaded_count
    
    def load_existing_granulat(self):
        """Lädt bestehende Granulat-Inventur"""
        if not self.inventur_granulat_path.exists():
            return 0
        
        loaded_count = 0
        try:
            # Lade Granulat Inventur-Sheet
            df_inventur = pd.read_excel(self.inventur_granulat_path, sheet_name='Inventur', dtype={'Charge': str})
            for _, row in df_inventur.iterrows():
                bemerkung = row.get('Bemerkung', '')
                if pd.isna(bemerkung) or str(bemerkung).lower() == 'nan':
                    bemerkung = ''
                
                item = {
                    'zeitstempel': str(row.get('Datum/Uhrzeit', '')),
                    'charge': str(row.get('Charge', '')),
                    'material': str(row.get('Material', '')),
                    'kurztext': str(row.get('Materialkurztext', '')),
                    'frei_verwendbar_kg': float(row.get('Frei verwendbar (KG)', 0)),
                    'zahlmenge_kg': float(row.get('Zählmenge (KG)', 0)),
                    'bemerkung': str(bemerkung),
                    'status': 'gefunden',
                    'typ': 'GRANULAT'
                }
                self.inventur_granulat_data.append(item)
                loaded_count += 1
            
            # Lade Granulat Nicht_gefunden-Sheet
            try:
                df_nicht_gefunden = pd.read_excel(self.inventur_granulat_path, sheet_name='Nicht_gefunden', dtype={'Charge': str})
                for _, row in df_nicht_gefunden.iterrows():
                    bemerkung = row.get('Bemerkung', '')
                    if pd.isna(bemerkung) or str(bemerkung).lower() == 'nan':
                        bemerkung = ''
                    
                    item = {
                        'zeitstempel': str(row.get('Datum/Uhrzeit', '')),
                        'charge': str(row.get('Charge', '')),
                        'material': str(row.get('Material', '')),
                        'kurztext': str(row.get('Materialkurztext', '')),
                        'frei_verwendbar_kg': float(row.get('Frei verwendbar (KG)', 0)),
                        'zahlmenge_kg': float(row.get('Zählmenge (KG)', 0)),
                        'bemerkung': str(bemerkung),
                        'status': 'nicht_gefunden',
                        'typ': 'GRANULAT'
                    }
                    self.nicht_gefunden_granulat_data.append(item)
                    loaded_count += 1
            except:
                pass  # Sheet existiert noch nicht
                
        except Exception as e:
            self.logger.error(f"Fehler beim Laden der Granulat-Inventur: {e}")
        
        return loaded_count
    
    def export_inventur(self):
        """Exportiert beide Inventur-Dateien als Backup (V2)"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # Erstelle Backup-Verzeichnis falls nicht vorhanden
            backup_dir = self.data_dir / 'backups'
            backup_dir.mkdir(exist_ok=True)
            
            # Backup-Dateinamen
            rollen_backup = backup_dir / f"Inventur_Rollen_Backup_{timestamp}.xlsx"
            granulat_backup = backup_dir / f"Inventur_Granulat_Backup_{timestamp}.xlsx"
            
            import shutil
            backup_count = 0
            
            # Kopiere Rollen-Datei falls vorhanden
            if self.inventur_rollen_path.exists():
                shutil.copy2(self.inventur_rollen_path, rollen_backup)
                backup_count += 1
            else:
                # Erstelle neue Rollen-Datei
                self.save_rollen_excel()
                if self.inventur_rollen_path.exists():
                    shutil.copy2(self.inventur_rollen_path, rollen_backup)
                    backup_count += 1
            
            # Kopiere Granulat-Datei falls vorhanden
            if self.inventur_granulat_path.exists():
                shutil.copy2(self.inventur_granulat_path, granulat_backup)
                backup_count += 1
            else:
                # Erstelle neue Granulat-Datei
                self.save_granulat_excel()
                if self.inventur_granulat_path.exists():
                    shutil.copy2(self.inventur_granulat_path, granulat_backup)
                    backup_count += 1
            
            if backup_count > 0:
                backup_message = f"Backup erfolgreich erstellt:\n\n"
                if rollen_backup.exists():
                    backup_message += f"🔵 Rollen: {rollen_backup.name}\n"
                if granulat_backup.exists():
                    backup_message += f"🟨 Granulat: {granulat_backup.name}\n"
                backup_message += f"\nSpeicherort: {backup_dir}"
                
                messagebox.showinfo("Export erfolgreich", backup_message)
                self.logger.info(f"V2 Backup erstellt: {backup_count} Dateien")
            else:
                messagebox.showwarning("Warnung", "Keine Inventur-Daten zum Exportieren gefunden.")
                
        except Exception as e:
            messagebox.showerror("Fehler", f"Fehler beim Export:\n{e}")
            self.logger.error(f"Fehler beim Export: {e}")
    
    def show_context_menu(self, event):
        """Zeigt Kontextmenü für Listeneinträge"""
        item = self.tree.selection()[0] if self.tree.selection() else None
        if not item:
            return
        
        context_menu = tk.Menu(self.root, tearoff=0)
        context_menu.add_command(label="🗑️ Löschen", command=lambda: self.delete_entry(item))
        # Bearbeitungsfunktion temporär entfernt
        # context_menu.add_command(label="✏️ Bearbeiten", command=lambda: self.edit_entry())
        
        try:
            context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            context_menu.grab_release()
    
    def delete_entry(self, item_id):
        """Löscht einen Eintrag (V2)"""
        if not messagebox.askyesno("Löschen bestätigen", 
                                  "Möchten Sie diesen Eintrag wirklich löschen?"):
            return
        
        # Finde Eintrag in Daten
        values = self.tree.item(item_id, 'values')
        if len(values) >= 2:
            charge = values[1]  # Charge ist in Spalte 1
            
            # Entferne aus allen Listen
            self.inventur_rollen_data = [item for item in self.inventur_rollen_data if item['charge'] != charge]
            self.nicht_gefunden_rollen_data = [item for item in self.nicht_gefunden_rollen_data if item['charge'] != charge]
            self.inventur_granulat_data = [item for item in self.inventur_granulat_data if item['charge'] != charge]
            self.nicht_gefunden_granulat_data = [item for item in self.nicht_gefunden_granulat_data if item['charge'] != charge]
            
            # Speichere und aktualisiere
            self.save_to_excel()
            self.update_list()
            
            self.status_var.set("Eintrag gelöscht")
            self.logger.info(f"Eintrag gelöscht: {charge}")
    
    def edit_entry(self):
        """Bearbeitet einen Eintrag (temporär deaktiviert)"""
        messagebox.showinfo("Info", "Bearbeitungsfunktion ist temporär deaktiviert.")
    
    # find_item_by_charge temporär entfernt (für Bearbeitungsfunktion)
    # def find_item_by_charge(self, charge):
    #     """Findet einen Eintrag anhand der Charge in allen Listen"""
    #     pass
    
    def undo_last_action(self):
        """Macht die letzte Aktion rückgängig (V2)"""
        if not self.undo_stack:
            self.status_var.set("Nichts zum Rückgängigmachen")
            return
        
        action, data, typ = self.undo_stack.pop()
        
        if action == 'add':
            # Entferne letzten Eintrag basierend auf Typ
            charge = data['charge']
            
            if typ == 'ROLLE':
                if data['status'] == 'gefunden':
                    self.inventur_rollen_data = [item for item in self.inventur_rollen_data if item['charge'] != charge]
                else:
                    self.nicht_gefunden_rollen_data = [item for item in self.nicht_gefunden_rollen_data if item['charge'] != charge]
            elif typ == 'GRANULAT':
                if data['status'] == 'gefunden':
                    self.inventur_granulat_data = [item for item in self.inventur_granulat_data if item['charge'] != charge]
                else:
                    self.nicht_gefunden_granulat_data = [item for item in self.nicht_gefunden_granulat_data if item['charge'] != charge]
            
            self.save_to_excel()
            self.update_list()
            
            typ_icon = "🔵" if typ == 'ROLLE' else "🟨"
            self.status_var.set(f"{typ_icon} Eintrag rückgängig gemacht: {charge}")
            self.logger.info(f"Undo {typ}: {charge}")
    
    def manual_save(self):
        """Manuelles Speichern"""
        self.save_to_excel()
        self.status_var.set("Manuell gespeichert")
    
    def toggle_fullscreen(self):
        """Schaltet Vollbild-Modus um"""
        current_state = self.root.attributes('-fullscreen')
        self.root.attributes('-fullscreen', not current_state)
        
        if not current_state:
            self.status_var.set("Vollbild-Modus aktiviert (F11 zum Beenden)")
        else:
            self.status_var.set("Vollbild-Modus deaktiviert")
    
    def ensure_scan_focus(self, event=None):
        """Stellt sicher, dass Scan-Feld fokussiert bleibt"""
        if hasattr(self, 'scan_entry') and not self.current_scan:
            self.root.after(100, lambda: self.scan_entry.focus_set())
    
    def quit_app(self):
        """Beendet die Anwendung"""
        if messagebox.askyesno("Beenden", "Möchten Sie das Programm wirklich beenden?"):
            self.logger.info("Programm beendet")
            self.root.quit()
    
    def run(self):
        """Startet die Anwendung"""
        self.root.mainloop()


# EditItemDialog temporär entfernt für Überarbeitung
# class EditItemDialog:
#     pass


class NotFoundDialog:
    """Dialog für nicht gefundene Waren (V2 mit Typ-Auswahl)"""
    
    def __init__(self, parent, charge):
        self.result = None
        self.selected_type = None
        self.input_widgets = {}
        
        # Dialog-Fenster
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("⚠️ Ware nicht gefunden - Typ wählen")
        self.dialog.geometry("550x750")  # Einheitliche Größe für alle Felder
        self.dialog.resizable(True, True)
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Zentriere Dialog
        self.dialog.geometry("+%d+%d" % (parent.winfo_rootx() + 50, parent.winfo_rooty() + 50))
        
        self.create_widgets(charge)
        
        # Warte auf Schließen
        self.dialog.wait_window()
    
    def create_widgets(self, charge):
        """Erstellt Dialog-Widgets mit Typ-Auswahl"""
        main_frame = ttk.Frame(self.dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Titel
        title_label = ttk.Label(main_frame, 
                               text=f"Ware mit Charge {charge} nicht gefunden!",
                               font=("Arial", 12, "bold"),
                               foreground="red")
        title_label.pack(pady=(0, 15))
        
        # Info
        info_label = ttk.Label(main_frame, 
                              text="Wählen Sie zuerst den Warentyp:",
                              font=("Arial", 10))
        info_label.pack(pady=(0, 10))
        
        # Typ-Auswahl Frame
        type_frame = ttk.LabelFrame(main_frame, text="Warentyp auswählen", padding="10")
        type_frame.pack(fill=tk.X, pady=(0, 20))
        
        self.type_var = tk.StringVar(value="ROLLE")  # Standard: Rolle
        
        # Rolle Radio Button
        rolle_radio = ttk.Radiobutton(type_frame, 
                                     text="🔵 Rolle (mit Länge, Breite, Fach)",
                                     variable=self.type_var, 
                                     value="ROLLE",
                                     command=self.on_type_change)
        rolle_radio.pack(anchor=tk.W, pady=2)
        
        # Granulat Radio Button
        granulat_radio = ttk.Radiobutton(type_frame, 
                                        text="🟨 Granulat (mit Gewicht)",
                                        variable=self.type_var, 
                                        value="GRANULAT",
                                        command=self.on_type_change)
        granulat_radio.pack(anchor=tk.W, pady=2)
        
        # Basis-Eingabefelder
        self.charge_var = tk.StringVar(value=charge)
        self.material_var = tk.StringVar()
        self.kurztext_var = tk.StringVar()
        self.bemerkung_var = tk.StringVar()
        
        # Basis-Felder erstellen
        self.create_base_fields(main_frame)
        
        # Container für dynamische Felder
        self.dynamic_frame = ttk.LabelFrame(main_frame, text="Spezifische Daten", padding="10")
        self.dynamic_frame.pack(fill=tk.X, pady=(10, 0))
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=(20, 10), fill=tk.X)
        
        # Zentriere Buttons
        inner_button_frame = ttk.Frame(button_frame)
        inner_button_frame.pack()
        
        save_button = ttk.Button(inner_button_frame, text="💾 Speichern", 
                                command=self.save_data, width=12)
        save_button.pack(side=tk.LEFT, padx=(0, 15))
        
        cancel_button = ttk.Button(inner_button_frame, text="❌ Abbrechen", 
                                  command=self.cancel, width=12)
        cancel_button.pack(side=tk.LEFT)
        
        # Enter-Binding
        self.dialog.bind('<Return>', lambda e: self.save_data())
        self.dialog.bind('<Escape>', lambda e: self.cancel())
        
        # Initial: Rolle-Felder anzeigen
        self.on_type_change()
    
    def create_base_fields(self, parent):
        """Erstellt die Basis-Eingabefelder"""
        base_frame = ttk.LabelFrame(parent, text="Grunddaten", padding="10")
        base_frame.pack(fill=tk.X, pady=(0, 10))
        
        base_fields = [
            ("Charge:", self.charge_var, False),
            ("Material-Nummer:", self.material_var, True),
            ("Materialkurztext:", self.kurztext_var, False),
            ("Bemerkung:", self.bemerkung_var, False)
        ]
        
        for i, (label_text, var, required) in enumerate(base_fields):
            field_frame = ttk.Frame(base_frame)
            field_frame.pack(fill=tk.X, pady=3)
            
            display_text = label_text + " *" if required else label_text
            label = ttk.Label(field_frame, text=display_text, font=("Arial", 9))
            label.pack(anchor=tk.W)
            
            entry = ttk.Entry(field_frame, textvariable=var, width=50, font=("Arial", 10))
            entry.pack(fill=tk.X, pady=(1, 0))
            
            if i == 1:  # Fokus auf Material-Nummer
                entry.focus_set()
    
    def on_type_change(self):
        """Wird aufgerufen wenn der Typ geändert wird"""
        # Lösche alte dynamische Felder
        for widget in self.dynamic_frame.winfo_children():
            widget.destroy()
        self.input_widgets.clear()
        
        selected_type = self.type_var.get()
        
        if selected_type == "ROLLE":
            self.create_rolle_fields()
        elif selected_type == "GRANULAT":
            self.create_granulat_fields()
    
    def create_rolle_fields(self):
        """Erstellt Eingabefelder für Rollen"""
        # Rollen-spezifische Variablen
        self.laenge_var = tk.StringVar()
        self.breite_var = tk.StringVar()
        self.flaeche_var = tk.StringVar()
        self.fach_var = tk.StringVar()
        
        rolle_fields = [
            ("Länge (m):", self.laenge_var, True),
            ("Breite (mm):", self.breite_var, True),
            ("Fläche (m²):", self.flaeche_var, True),
            ("Fach:", self.fach_var, True)
        ]
        
        for label_text, var, required in rolle_fields:
            field_frame = ttk.Frame(self.dynamic_frame)
            field_frame.pack(fill=tk.X, pady=3)
            
            display_text = label_text + " *" if required else label_text
            label = ttk.Label(field_frame, text=display_text, font=("Arial", 9))
            label.pack(anchor=tk.W)
            
            entry = ttk.Entry(field_frame, textvariable=var, width=50, font=("Arial", 10))
            entry.pack(fill=tk.X, pady=(1, 0))
            
            self.input_widgets[label_text] = entry
    
    def create_granulat_fields(self):
        """Erstellt Eingabefelder für Granulat"""
        # Granulat-spezifische Variablen
        self.frei_verwendbar_var = tk.StringVar()
        
        granulat_fields = [
            ("Frei verwendbar (KG):", self.frei_verwendbar_var, True)
        ]
        
        for label_text, var, required in granulat_fields:
            field_frame = ttk.Frame(self.dynamic_frame)
            field_frame.pack(fill=tk.X, pady=3)
            
            display_text = label_text + " *" if required else label_text
            label = ttk.Label(field_frame, text=display_text, font=("Arial", 9))
            label.pack(anchor=tk.W)
            
            entry = ttk.Entry(field_frame, textvariable=var, width=50, font=("Arial", 10))
            entry.pack(fill=tk.X, pady=(1, 0))
            
            self.input_widgets[label_text] = entry
    
    def save_data(self):
        """Speichert die eingegebenen Daten basierend auf dem gewählten Typ"""
        selected_type = self.type_var.get()
        
        # Basis-Validierung
        if not self.material_var.get().strip():
            messagebox.showerror("Fehler", "Bitte Material-Nummer eingeben")
            return
        
        try:
            # Basis-Daten
            result_data = {
                'charge': self.charge_var.get().strip(),
                'material': self.material_var.get().strip(),
                'kurztext': self.kurztext_var.get().strip(),
                'bemerkung': self.bemerkung_var.get().strip(),
                'typ': selected_type,
                'status': 'nicht_gefunden'
            }
            
            if selected_type == "ROLLE":
                # Rollen-Validierung
                required_rolle_fields = [
                    (self.laenge_var.get().strip(), "Länge"),
                    (self.breite_var.get().strip(), "Breite"),
                    (self.flaeche_var.get().strip(), "Fläche"),
                    (self.fach_var.get().strip(), "Fach")
                ]
                
                for value, field_name in required_rolle_fields:
                    if not value:
                        messagebox.showerror("Fehler", f"Bitte {field_name} eingeben")
                        return
                
                # Konvertiere Rollen-Werte
                # Für nicht gefundene Rollen: Alle Werte sind manuell eingegeben
                breite_eingegeben = int(self.breite_var.get())
                result_data.update({
                    'laenge': float(self.laenge_var.get().replace(',', '.')),
                    'breite_original': breite_eingegeben,  # Manuell eingegeben
                    'breite_kontrolliert': breite_eingegeben,  # Gleicher Wert, da manuell
                    'flaeche': float(self.flaeche_var.get().replace(',', '.')),
                    'fach_original': '',  # Leer, da nicht aus Arbeitstabelle
                    'fach_kontrolliert': self.fach_var.get().strip()  # Manuell eingegeben
                })
                
            elif selected_type == "GRANULAT":
                # Granulat-Validierung
                if not self.frei_verwendbar_var.get().strip():
                    messagebox.showerror("Fehler", "Bitte Frei verwendbar (KG) eingeben")
                    return
                
                # Konvertiere Granulat-Werte
                result_data.update({
                    'frei_verwendbar_kg': float(self.frei_verwendbar_var.get().replace(',', '.'))
                })
            
            self.result = result_data
            self.dialog.destroy()
            
        except ValueError:
            messagebox.showerror("Fehler", "Ungültige Zahlenwerte eingegeben")
    
    def cancel(self):
        """Bricht den Dialog ab"""
        self.result = None
        self.dialog.destroy()


def main():
    """Hauptfunktion"""
    try:
        app = InventurApp()
        app.run()
    except Exception as e:
        messagebox.showerror("Kritischer Fehler", f"Unerwarteter Fehler:\n{e}")
        logging.error(f"Kritischer Fehler: {e}")


if __name__ == "__main__":
    main()
